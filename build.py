#!/usr/bin/env python3
"""Build site/data.json from the Excel inventory files in data/.

Each .xlsx in data/ is one equipment category (filename = category key in
config.json). Sheets must follow the standard INC.xlsx column layout:

    الماركة | الموديل | الرقم المسلسل | الكود | تاريخ الانتاج | TRC | الحالة الفنية | وصف الحالة الفنية
"""
import json
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
OUT_FILE = ROOT / "site" / "data.json"
MAINT_LOG = DATA_DIR / "maintenance_log.xlsx"

CURRENT_YEAR = datetime.now().year
TODAY = date.today()

# Column header -> field name (headers as they appear in the standard sheet)
HEADER_MAP = {
    "الماركة": "brand",
    "الموديل": "model",
    "الرقم المسلسل": "serial",
    "الكود": "code",
    "تاريخ الانتاج": "year",
    "TRC": "trc",
    "شهادة الجودة": "certified",
    "الحالة الفنية": "status",
    "وصف الحالة الفنية": "description",
}

STATUS_RANK = {"NF": 3, "PF": 2, "FF": 1}

CERTIFIED_YES = {"yes", "نعم", "معتمد", "true", "1"}

# Manufacturer support status (Asset Management Classification document).
# Canonical keys used internally; display names + definitions exported to data.json.
SUPPORT_MAP = {
    "supported/active": "supported", "supported": "supported", "active": "supported",
    "limited supported": "limited", "limited support": "limited",
    "limited": "limited", "legacy": "limited",
    "eol": "eol", "end of life": "eol",
    "obsolete": "obsolete", "obsolet": "obsolete", "obsulet": "obsolete",
}

ASSET_CLASSIFICATION = {
    "supported": {
        "label_en": "Supported / Active", "label_ar": "مدعوم / نشط",
        "def_en": "Manufacturer provides technical support, software updates, training, and spare parts.",
        "def_ar": "الشركة المصنعة توفر الدعم الفني والتحديثات والتدريب وقطع الغيار.",
    },
    "limited": {
        "label_en": "Limited Support (Legacy)", "label_ar": "دعم محدود",
        "def_en": "No longer actively marketed but some spare parts and service support remain available.",
        "def_ar": "لم يعد يُسوَّق ولكن بعض قطع الغيار وخدمات الدعم لا تزال متاحة.",
    },
    "eol": {
        "label_en": "End of Life (EOL)", "label_ar": "نهاية العمر التشغيلي",
        "def_en": "Manufacturer has ended routine support; parts may be difficult to obtain.",
        "def_ar": "أنهت الشركة المصنعة الدعم الدوري؛ قد يصعب الحصول على قطع الغيار.",
    },
    "obsolete": {
        "label_en": "Obsolete", "label_ar": "متقادم",
        "def_en": "Should be considered for replacement due to age, support limitations, technology advancement, or regulatory concerns.",
        "def_ar": "يجب النظر في استبداله بسبب العمر أو محدودية الدعم أو تقدم التقنية أو الاعتبارات التنظيمية.",
    },
}


def parse_support(v):
    return SUPPORT_MAP.get(clean(v).lower())

# Spare-part patterns, most specific first. Each: (part_key, regex)
PART_PATTERNS = [
    ("humidifier_door_lock", r"لوك\s+باب\s+المرطب"),
    ("humidifier_door", r"باب\s+المرطب"),
    ("humidifier", r"المرطب"),
    ("window_lock", r"لوك\s+شباك"),
    ("door_lock", r"لوك\s+باب"),
    ("window", r"شباك"),
    ("hood", r"hood"),
    ("door", r"باب|الابواب|الأبواب"),
    ("control_board", r"بوردة\s+الكنترول|بورده\s+الكنترول"),
    ("motor", r"الماتور|ماتور"),
    ("heater", r"heater|الهيتر"),
    ("keypad", r"stuck\s*key"),
    ("monitor_mount", r"حامل\s+الشاشة"),
]

PART_NAMES = {
    "window":               {"en": "Porthole window",        "ar": "شباك"},
    "window_lock":          {"en": "Porthole window lock",   "ar": "لوك شباك"},
    "door":                 {"en": "Access door",            "ar": "باب"},
    "door_lock":            {"en": "Access door lock",       "ar": "لوك باب"},
    "humidifier":           {"en": "Humidifier unit",        "ar": "وحدة الترطيب (المرطب)"},
    "humidifier_door":      {"en": "Humidifier door",        "ar": "باب المرطب"},
    "humidifier_door_lock": {"en": "Humidifier door lock",   "ar": "لوك باب المرطب"},
    "hood":                 {"en": "Hood (canopy)",          "ar": "الغطاء (hood)"},
    "control_board":        {"en": "Control board",          "ar": "بوردة الكنترول"},
    "motor":                {"en": "Motor",                  "ar": "الماتور"},
    "heater":               {"en": "Heater",                 "ar": "الهيتر (heater)"},
    "keypad":               {"en": "Keypad / membrane keys", "ar": "لوحة المفاتيح"},
    "monitor_mount":        {"en": "Monitor mount fixing",   "ar": "حامل الشاشة"},
}

# TRC criteria from the organization's classification document (DOC-20260501-WA0011)
TRC_CRITERIA = [
    {
        "trc": 1,
        "verdict_ar": "يعمل ويعتمد عليه",
        "verdict_en": "Working and dependable",
        "criteria_ar": [
            "أحدث إنتاج للشركة المصنعة ولم يمر على إنتاجه أكثر من 3 سنوات ويعمل بكفاءة وحاصل على شهادة جودة معتمدة",
        ],
        "criteria_en": [
            "Latest production from the manufacturer, less than 3 years old, working efficiently, with a certified quality certificate",
        ],
    },
    {
        "trc": 2,
        "verdict_ar": "يعمل ويعتمد عليه",
        "verdict_en": "Working and dependable",
        "criteria_ar": [
            "حديث نسبياً لم يمر على إنتاجه أكثر من 7 سنوات",
            "يوجد له دعم فني من الشركة المصنعة والوكيل المعتمد",
            "أعطال الجهاز لا تتخطى 20% من السنة والتكلفة التراكمية لا تزيد عن 20% عن الجهاز المماثل",
            "الجهاز يستخدم بصورة أساسية ويعتمد عليه",
        ],
        "criteria_en": [
            "Relatively recent, less than 7 years since production",
            "Technical support available from manufacturer and authorized agent",
            "Downtime under 20% per year; cumulative cost under 20% of an equivalent device",
            "Used as primary equipment and relied upon",
        ],
    },
    {
        "trc": 3,
        "verdict_ar": "يعمل ويعتمد عليه",
        "verdict_en": "Working and dependable",
        "criteria_ar": [
            "توقف إنتاجه ولكن يوجد له دعم فني من الشركة المصنعة والوكيل المعتمد",
            "التكلفة التراكمية للجهاز لا تزيد عن 50% من سعر الجهاز",
            "نسبة الأعطال (Down time) لا تتخطى الـ 50%",
        ],
        "criteria_en": [
            "Production discontinued, but technical support still available from manufacturer and agent",
            "Cumulative cost under 50% of the device price",
            "Downtime does not exceed 50%",
        ],
    },
    {
        "trc": 4,
        "verdict_ar": "يعمل ولكن لا يعتمد عليه بصورة أساسية",
        "verdict_en": "Working but not dependable as primary equipment",
        "criteria_ar": [
            "توقف إنتاجه ولا يوجد له دعم فني أو قطع غيار من الشركة المصنعة والوكيل المعتمد",
            "لا يمكن الاعتماد عليه بصورة أساسية",
            "كثير الأعطال ونسبة الأعطال (Down time) تتخطى الـ 50%",
            "التكلفة التراكمية للجهاز تزيد عن 50% من سعر الجهاز",
        ],
        "criteria_en": [
            "Production discontinued; no technical support or spare parts from manufacturer or agent",
            "Cannot be relied upon as primary equipment",
            "Frequent failures; downtime exceeds 50%",
            "Cumulative cost exceeds 50% of the device price",
        ],
    },
    {
        "trc": 5,
        "verdict_ar": "معطل",
        "verdict_en": "Out of service",
        "criteria_ar": [
            "غير آمن",
            "لا جدوى فنية أو اقتصادية من إصلاحه",
            "مستلزمات التشغيل أو قطع الغيار والأكسسوارات الأساسية غير متوفرة",
        ],
        "criteria_en": [
            "Unsafe",
            "No technical or economic feasibility of repair",
            "Operating supplies, spare parts, and essential accessories unavailable",
        ],
    },
]


def clean(v):
    if v is None:
        return ""
    return str(v).strip()


def parse_year(v):
    m = re.search(r"(19|20)\d{2}", clean(v))
    return int(m.group(0)) if m else None


def parse_faults(description):
    """Split multi-line fault description into individual fault strings."""
    lines = [ln.strip() for ln in clean(description).splitlines()]
    return [ln for ln in lines if ln]


def extract_parts(fault_line):
    """Detect needed spare parts (with quantity) in one Arabic fault line."""
    found = []
    remaining = fault_line
    for key, pattern in PART_PATTERNS:
        m = re.search(pattern, remaining, re.IGNORECASE)
        if not m:
            continue
        qty = 1
        # quantity forms: "عدد (2) شباك", "عدد (2)شباك", "2 شباك", "تحتاج 3 شباك"
        prefix = remaining[: m.start()]
        qm = re.search(r"عدد\s*\(?\s*(\d+)\s*\)?\s*$", prefix) or re.search(r"(\d+)\s*$", prefix)
        if qm:
            qty = int(qm.group(1))
        found.append({"part": key, "qty": qty})
        # remove the matched text so broader patterns (شباك after لوك شباك) don't double-count
        remaining = remaining[: m.start()] + remaining[m.end():]
    return found


def parse_log_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = clean(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


LOG_HEADER_MAP = {
    "التاريخ": "date",
    "الكود": "code",
    "حالة الدعم الفني": "support",
    "نوع العمل": "wtype",
    "الوصف": "desc",
    "أيام التوقف": "downtime",
    "التكلفة": "cost",
}


def read_maintenance_log():
    """Read data/maintenance_log.xlsx → ({device_code: [event, ...]}, {device_code: support}).

    Columns: التاريخ | الكود | حالة الدعم الفني | نوع العمل | الوصف | أيام التوقف | التكلفة
    (parsed by header name, so the support column is optional in older files).
    نوع العمل containing 'وقائ' (or PM) is preventive; anything else is a repair.
    حالة الدعم الفني: latest non-empty value per device wins.
    """
    events, support = {}, {}
    if not MAINT_LOG.exists():
        return events, support
    ws = openpyxl.load_workbook(MAINT_LOG, data_only=True)["Log"]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows, [])]
    fields = [LOG_HEADER_MAP.get(h) for h in headers]
    for row in rows:
        rec = {f: v for f, v in zip(fields, row) if f}
        code = clean(rec.get("code"))
        d = parse_log_date(rec.get("date"))
        if not code or not d:
            continue
        kind = "pm" if re.search(r"وقائ|pm", clean(rec.get("wtype")), re.IGNORECASE) else "repair"
        events.setdefault(code, []).append({
            "date": d.isoformat(),
            "type": kind,
            "desc": clean(rec.get("desc")),
            "downtime": float(rec["downtime"]) if clean(rec.get("downtime")) not in ("", "N/A") else 0.0,
            "cost": float(rec["cost"]) if clean(rec.get("cost")) not in ("", "N/A") else 0.0,
            "_support": parse_support(rec.get("support")),
        })
    for code, evs in events.items():
        evs.sort(key=lambda e: e["date"])
        for e in evs:
            s = e.pop("_support", None)
            if s:
                support[code] = s  # events are date-sorted, so the latest non-empty wins
    return events, support


def maintenance_analytics(evs, pm_interval_days):
    """Per-device maintenance KPIs from its log events."""
    if not evs:
        return None
    year_ago = (TODAY - timedelta(days=365)).isoformat()
    repairs = [e for e in evs if e["type"] == "repair"]
    pms = [e for e in evs if e["type"] == "pm"]
    recent = [e for e in evs if e["date"] >= year_ago]
    recent_repairs = [e for e in recent if e["type"] == "repair"]

    downtime_12m = sum(e["downtime"] for e in recent_repairs)
    mtbf = None
    if len(repairs) >= 2:
        dates = [date.fromisoformat(e["date"]) for e in repairs]
        gaps = [(b - a).days for a, b in zip(dates, dates[1:])]
        mtbf = round(sum(gaps) / len(gaps))
    mttr = round(sum(e["downtime"] for e in repairs) / len(repairs), 1) if repairs else None

    last_pm = pms[-1]["date"] if pms else None
    next_pm_due = None
    pm_overdue = False
    if last_pm:
        due = date.fromisoformat(last_pm) + timedelta(days=pm_interval_days)
        next_pm_due = due.isoformat()
        pm_overdue = due < TODAY

    return {
        "events": evs[-15:],
        "repairs_12m": len(recent_repairs),
        "downtime_12m": round(downtime_12m, 1),
        "downtime_pct": round(100 * downtime_12m / 365, 1),
        "cost_12m": round(sum(e["cost"] for e in recent), 2),
        "cum_cost": round(sum(e["cost"] for e in evs), 2),
        "mtbf_days": mtbf,
        "mttr_days": mttr,
        "last_pm": last_pm,
        "next_pm_due": next_pm_due,
        "pm_overdue": pm_overdue,
    }


def get_expected_life(brand, model, cat_cfg, default_life):
    """Expected life: model override (BEAG) > category default > global default."""
    full = f"{brand} {model}".strip().lower()
    for key, years in cat_cfg.get("model_life_years", {}).items():
        if key.lower() == full or key.lower() in full:
            return years
    return cat_cfg.get("expected_life_years", default_life)


def model_support_status(brand, model, cat_cfg):
    """Config-level support default per model (overridden by maintenance-log entries)."""
    full = f"{brand} {model}".strip().lower()
    for key, value in cat_cfg.get("model_support_status", {}).items():
        if key.lower() == full or key.lower() in full:
            return parse_support(value)
    return None


def suggest_trc(age, status, maint, price, certified=False, support=None):
    """Data-driven TRC suggestion applying DOC-20260501-WA0011 criteria as AND-logic.

    TRC 1: age ≤ 3 AND fully functional AND certified AND manufacturer support active.
    TRC 2: age ≤ 7 AND downtime < 20% AND cumulative cost < 20% of price AND supported.
    TRC 3: discontinued but still supported (incl. limited/EOL), downtime ≤ 50%, cost ≤ 50%.
    TRC 4: downtime > 50% OR cost > 50% OR support Obsolete OR device is NF.
    TRC 5 is never auto-suggested — the unsafe / no-spare-parts judgment stays
    with the engineer (NF devices get TRC 4 + a review warning instead).

    Works without maintenance history: downtime/cost then count as 0 but are
    flagged unverified. Unknown support is assumed active, with a warning.
    Warnings are language-neutral codes translated by the frontend.
    """
    downtime_pct = maint["downtime_pct"] if maint else None
    cost_ratio = round(maint["cum_cost"] / price, 2) if (maint and price) else None
    warnings = []

    if maint is None:
        warnings.append("no_history")
    if support is None:
        support = "supported"
        warnings.append("support_unknown")
    if age is None:
        warnings.append("year_unknown")

    dt = downtime_pct or 0.0
    cr = cost_ratio or 0.0

    if dt > 50 or cr > 0.5 or support == "obsolete" or status == "NF":
        trc = 4
        if dt > 50:
            warnings.append("high_downtime")
        if cr > 0.5:
            warnings.append("high_cost")
        if support == "obsolete":
            warnings.append("obsolete")
        if status == "NF":
            warnings.append("nf_review")  # engineer should assess whether this is TRC 5
    elif (age is not None and age <= 3 and status == "FF"
          and certified and support == "supported"):
        trc = 1
    elif (age is not None and age <= 7 and dt < 20 and cr < 0.2
          and support == "supported"):
        trc = 2
        if age <= 3 and status == "FF" and not certified:
            warnings.append("cert_missing")  # would qualify for TRC 1 with a quality certificate
    else:
        trc = 3
        if age is not None and age > 7:
            warnings.append("age_over_7")
        if 20 <= dt <= 50:
            warnings.append("downtime_over_20")
        if 0.2 <= cr <= 0.5:
            warnings.append("cost_over_20")
        if support == "limited":
            warnings.append("limited_support")
        if support == "eol":
            warnings.append("eol")

    return {
        "trc": trc,
        "downtime_pct": downtime_pct,
        "cost_ratio": cost_ratio,
        "certified": certified,
        "support_status": support,
        "unverified": maint is None,
        "warnings": warnings,
    }


def score_device(trc, status, age, expected_life):
    """Explainable 0-100 replacement priority score."""
    trc = trc or 3
    score = trc * 15
    score += {"NF": 20, "PF": 8}.get(status, 0)
    if age is not None and expected_life:
        score += min(age / expected_life, 1.5) * 10
    score = min(round(score), 100)

    if trc == 5 or score >= 80:
        rec = "replace_now"
    elif score >= 65:
        rec = "plan_replacement"
    elif score >= 40:
        rec = "monitor"
    else:
        rec = "ok"
    return score, rec


def read_category(path, cat_cfg, default_life, log_events, log_support, config):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    fields = [HEADER_MAP.get(h) for h in headers]
    if "code" not in fields:
        sys.exit(f"ERROR: {path.name}: sheet does not follow the standard column layout "
                 f"(expected Arabic headers like {list(HEADER_MAP)})")

    devices = {}
    for row in rows:
        rec = {f: clean(v) for f, v in zip(fields, row) if f}
        if not any(rec.values()):
            continue
        code = rec.get("code") or f"?{len(devices)+1}"
        year = parse_year(rec.get("year"))
        trc = int(m.group(0)) if (m := re.search(r"[1-5]", rec.get("trc", ""))) else None
        status = rec.get("status", "").upper()
        certified = rec.get("certified", "").lower() in CERTIFIED_YES
        faults = parse_faults(rec.get("description"))

        if code in devices:
            # same device listed twice: merge faults, keep worst TRC/status
            d = devices[code]
            d["faults"] += [f for f in faults if f not in d["faults"]]
            if trc and trc > (d["trc"] or 0):
                d["trc"] = trc
            if STATUS_RANK.get(status, 0) > STATUS_RANK.get(d["status"], 0):
                d["status"] = status
            d["certified"] = d["certified"] or certified
            continue

        devices[code] = {
            "code": code,
            "brand": rec.get("brand", ""),
            "model": rec.get("model", ""),
            "serial": rec.get("serial", ""),
            "year": year,
            "trc": trc,
            "certified": certified,
            "status": status,
            "faults": faults,
        }

    pm_interval = cat_cfg.get("pm_interval_days", config.get("default_pm_interval_days", 90))
    price = cat_cfg.get("replacement_price")

    result = []
    for d in devices.values():
        age = (CURRENT_YEAR - d["year"]) if d["year"] else None
        d["age"] = age
        expected_life = get_expected_life(d["brand"], d["model"], cat_cfg, default_life)
        d["expected_life"] = expected_life
        parts = []
        for fl in d["faults"]:
            parts += extract_parts(fl)
        d["parts_needed"] = parts
        d["score"], d["recommendation"] = score_device(d["trc"], d["status"], age, expected_life)
        d["maint"] = maintenance_analytics(log_events.get(d["code"]), pm_interval)
        # support: latest maintenance-log entry wins, else config per-model default
        support = log_support.get(d["code"]) or model_support_status(d["brand"], d["model"], cat_cfg)
        suggestion = suggest_trc(age, d["status"], d["maint"], price,
                                 certified=d["certified"], support=support)
        d["trc_suggested"] = suggestion
        d["trc_mismatch"] = bool(suggestion and d["trc"] and suggestion["trc"] != d["trc"])
        result.append(d)

    result.sort(key=lambda d: -d["score"])
    return result


def category_stats(devices):
    n = len(devices)
    by_status = {"FF": 0, "PF": 0, "NF": 0}
    by_trc = {}
    by_rec = {"replace_now": 0, "plan_replacement": 0, "monitor": 0, "ok": 0}
    ages = []
    past_life = 0
    for d in devices:
        if d["status"] in by_status:
            by_status[d["status"]] += 1
        if d["trc"]:
            by_trc[str(d["trc"])] = by_trc.get(str(d["trc"]), 0) + 1
        by_rec[d["recommendation"]] += 1
        if d["age"] is not None:
            ages.append(d["age"])
            if d["age"] >= d["expected_life"]:  # per-device (model-based) life
                past_life += 1
    parts_total = {}
    for d in devices:
        for p in d["parts_needed"]:
            e = parts_total.setdefault(p["part"], {"qty": 0, "devices": []})
            e["qty"] += p["qty"]
            if d["code"] not in e["devices"]:
                e["devices"].append(d["code"])

    logged = [d for d in devices if d["maint"]]
    pm_logged = [d for d in logged if d["maint"]["last_pm"]]
    soon = (TODAY + timedelta(days=14)).isoformat()
    maint_stats = {
        "logged_devices": len(logged),
        "pm_logged_devices": len(pm_logged),
        "pm_overdue": sum(1 for d in pm_logged if d["maint"]["pm_overdue"]),
        "pm_due_soon": sum(1 for d in pm_logged
                           if not d["maint"]["pm_overdue"] and d["maint"]["next_pm_due"] <= soon),
        "pm_compliance_pct": round(100 * sum(1 for d in pm_logged if not d["maint"]["pm_overdue"])
                                   / len(pm_logged)) if pm_logged else None,
        "repairs_12m": sum(d["maint"]["repairs_12m"] for d in logged),
        "downtime_days_12m": round(sum(d["maint"]["downtime_12m"] for d in logged), 1),
        "cost_12m": round(sum(d["maint"]["cost_12m"] for d in logged), 2),
        "trc_mismatches": sum(1 for d in devices if d["trc_mismatch"]),
    }

    return {
        "total": n,
        "by_status": by_status,
        "by_trc": dict(sorted(by_trc.items())),
        "by_recommendation": by_rec,
        "avg_age": round(sum(ages) / len(ages), 1) if ages else None,
        "past_expected_life": past_life,
        "past_expected_life_pct": round(100 * past_life / len(ages)) if ages else None,
        "parts_needed": dict(sorted(parts_total.items(), key=lambda kv: -kv[1]["qty"])),
        "maintenance": maint_stats,
    }


def main():
    config = json.loads((ROOT / "config.json").read_text(encoding="utf-8"))
    default_life = config.get("default_expected_life_years", 10)
    log_events, log_support = read_maintenance_log()
    categories = {}
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        if path.name == MAINT_LOG.name or path.name.startswith("~$"):
            continue
        key = path.stem.lower()
        cat_cfg = config["categories"].get(key, {})
        devices = read_category(path, cat_cfg, default_life, log_events, log_support, config)
        stats = category_stats(devices)
        categories[key] = {
            "name_en": cat_cfg.get("name_en", key.title()),
            "name_ar": cat_cfg.get("name_ar", key),
            "expected_life_years": cat_cfg.get("expected_life_years", default_life),
            "life_note_en": cat_cfg.get("life_note_en", ""),
            "life_note_ar": cat_cfg.get("life_note_ar", ""),
            "stats": stats,
            "devices": devices,
        }
        print(f"{path.name}: {stats['total']} devices "
              f"(FF {stats['by_status']['FF']} / PF {stats['by_status']['PF']} / NF {stats['by_status']['NF']})")

    out = {
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "current_year": CURRENT_YEAR,
        "actions_url": config.get("repo_actions_url", ""),
        "part_names": PART_NAMES,
        "trc_criteria": TRC_CRITERIA,
        "asset_classification": ASSET_CLASSIFICATION,
        "categories": categories,
    }
    OUT_FILE.parent.mkdir(exist_ok=True)
    OUT_FILE.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"Wrote {OUT_FILE} ({OUT_FILE.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
